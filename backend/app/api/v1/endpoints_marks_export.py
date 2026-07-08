"""
SPARSH - Marks Export Endpoint

GET /marks/export?academic_year_id=&class_level_id=&section=&exam_id=

Streams an .xlsx file containing all marks for the specified
class/section/exam/year combination.

Columns: Roll No | Admission No | Student Name | [Subject Obtained] | [Subject Max] ...
One row per student, subjects ordered by display_order then name.
Students ordered by roll_number.

Authorization: All authenticated users (read-only).
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import get_db, get_current_active_user
from app.models.user import User
from app.models.student import Student, StudentEnrollment
from app.models.academic import AcademicYear, ClassLevel, Exam, Subject
from app.models.result import StudentResult

router = APIRouter()

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_ALT_FILL    = PatternFill("solid", fgColor="EFF6FF")
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@router.get("/export")
def export_marks(
    academic_year_id: int = Query(...),
    class_level_id: int = Query(...),
    section: str = Query(...),
    exam_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Export marks for a class/section/exam as a downloadable .xlsx file.
    """
    year = db.query(AcademicYear).filter_by(id=academic_year_id).first()
    cls  = db.query(ClassLevel).filter_by(id=class_level_id).first()
    exam = db.query(Exam).filter_by(id=exam_id).first()
    if not year: raise HTTPException(404, "Academic year not found.")
    if not cls:  raise HTTPException(404, "Class level not found.")
    if not exam: raise HTTPException(404, "Exam not found.")

    # Fetch all enrollments for this class/section/year ordered by roll_number
    enrollments = (
        db.query(StudentEnrollment)
        .filter_by(
            academic_year_id=academic_year_id,
            class_level_id=class_level_id,
            section=section.upper(),
        )
        .order_by(StudentEnrollment.roll_number)
        .all()
    )

    if not enrollments:
        raise HTTPException(404, "No students enrolled for this class/section/year.")

    enrollment_ids = [e.id for e in enrollments]

    # Fetch all results for these enrollments + this exam
    results = (
        db.query(StudentResult)
        .filter(
            StudentResult.student_enrollment_id.in_(enrollment_ids),
            StudentResult.exam_id == exam_id,
        )
        .all()
    )

    # Get ordered list of subjects from results
    subject_ids_seen = list(dict.fromkeys(r.subject_id for r in results))
    subjects = (
        db.query(Subject)
        .filter(Subject.id.in_(subject_ids_seen))
        .order_by(Subject.subject_name)
        .all()
    ) if subject_ids_seen else db.query(Subject).order_by(Subject.subject_name).all()

    # Build result lookup: enrollment_id → subject_id → result
    result_map: dict[int, dict[int, StudentResult]] = {}
    for r in results:
        result_map.setdefault(r.student_enrollment_id, {})[r.subject_id] = r

    # Build student name lookup
    adm_numbers = [e.admission_number for e in enrollments]
    students = db.query(Student).filter(Student.admission_number.in_(adm_numbers)).all()
    name_map = {s.admission_number: s.student_name for s in students}

    # ── Build xlsx ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{cls.class_name} {section.upper()} {exam.exam_name}"

    # Headers
    fixed_headers = ["Roll No", "Admission No", "Student Name"]
    subject_headers = []
    for subj in subjects:
        subject_headers += [f"{subj.subject_name} (Obt)", f"{subj.subject_name} (Max)"]

    all_headers = fixed_headers + subject_headers
    ws.row_dimensions[1].height = 26

    for col_idx, hdr in enumerate(all_headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
        cell = ws.cell(1, col_idx, hdr)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    # Data rows
    for row_idx, enroll in enumerate(enrollments, 2):
        row_data = [
            enroll.roll_number or "",
            enroll.admission_number,
            name_map.get(enroll.admission_number, "—"),
        ]
        for subj in subjects:
            res = result_map.get(enroll.id, {}).get(subj.id)
            row_data.append(res.marks_obtained if res and res.marks_obtained is not None else "")
            row_data.append(res.max_marks      if res and res.max_marks      is not None else "")

        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, val)
            if fill: cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    ws.freeze_panes = "D2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"SPARSH_{cls.class_name}_{section.upper()}_{exam.exam_name}_{year.year_label}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
