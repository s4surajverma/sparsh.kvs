"""
SPARSH - Excel Template Download Endpoints

Generates .xlsx template files with header + sample row + notes row.

Template types:
  student_master  — admission_number, student_name, class_name, section, roll_number
  results         — admission_number, student_name, roll_number + one col per Subject in DB

Authorization: All authenticated users.
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import get_db, get_current_active_user
from app.models.user import User
from app.models.academic import Subject

router_templates_dl = APIRouter()

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_SAMPLE_FILL = PatternFill("solid", fgColor="EFF6FF")
_SAMPLE_FONT = Font(name="Calibri", color="1E3A5F", size=10)
_NOTE_FILL  = PatternFill("solid", fgColor="FFF9C4")
_NOTE_FONT  = Font(name="Calibri", italic=True, color="555555", size=9)
_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L = Alignment(horizontal="left",   vertical="center")
_THIN    = Side(style="thin", color="CCCCCC")
_BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _h(cell):
    cell.fill = _HEADER_FILL; cell.font = _HEADER_FONT
    cell.alignment = _ALIGN_C; cell.border = _BORDER

def _s(cell):
    cell.fill = _SAMPLE_FILL; cell.font = _SAMPLE_FONT
    cell.alignment = _ALIGN_L; cell.border = _BORDER

def _n(cell):
    cell.fill = _NOTE_FILL; cell.font = _NOTE_FONT
    cell.alignment = _ALIGN_L; cell.border = _BORDER


def _stream(wb, filename):
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _write_template(ws, cols):
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 32
    for i, (header, sample, note) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
        _h(ws.cell(1, i, header))
        _s(ws.cell(2, i, sample))
        _n(ws.cell(3, i, note))


@router_templates_dl.get("/download/student_master")
def download_student_master_template(
    current_user: User = Depends(get_current_active_user),
):
    """Download Student Master import template (.xlsx) with header + sample + notes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Master"
    _write_template(ws, [
        ("admission_number", "ADM001",         "Unique admission number (required)"),
        ("student_name",     "Sample Student",  "Full name of student (required)"),
        ("class_name",       "Class 6",         "Must match an existing Class Level exactly"),
        ("section",          "A",               "Section letter (e.g. A, B)"),
        ("roll_number",      "1",               "Roll number (numeric)"),
    ])
    ws.freeze_panes = "A2"
    return _stream(wb, "SPARSH_Student_Master_Template.xlsx")


@router_templates_dl.get("/download/results")
def download_results_template(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Download Results import template (.xlsx). Subject columns populated from DB."""
    subjects = db.query(Subject).order_by(Subject.subject_name).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    cols = [
        ("admission_number", "ADM001",        "Unique admission number (required)"),
        ("student_name",     "Sample Student", "Student full name"),
        ("roll_number",      "1",              "Roll number (numeric)"),
    ] + [
        (subj.subject_name, "75", f"Marks obtained in {subj.subject_name}")
        for subj in subjects
    ]

    _write_template(ws, cols)
    ws.freeze_panes = "D2"
    return _stream(wb, "SPARSH_Results_Template.xlsx")
